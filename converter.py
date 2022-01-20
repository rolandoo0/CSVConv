import csv
import os
import sys
import pandas as pd
import datetime
from pandas import ExcelWriter
from openpyxl import load_workbook
import time
import colorama
from colorama import Fore, Back, Style
import codecs

columnNames = [
    "order-id",
    "buyer-phone-number",
    "product-name",
    "quantity-to-ship",
    "recipient-name",
    "ship-address-1",
    "ship-address-2",
    "ship-city",
    "ship-state",
    "ship-postal-code",
]

def convertAFTERQUESTION(input_file):
    with open(input_file, "r", newline='') as csv_in_file:
        with open("output.csv", "w", newline="") as csv_out_file:
            filereader = csv.reader(csv_in_file)
            filewriter = csv.writer(csv_out_file)
            try:
                for row in filereader:
                    newRow = []
                    length = len(row)
                    if length > 1:
                        string = ""
                        for i in range(length):
                            string = string + row[i]
                        rows = string.split("\t")
                        for subRow in rows:
                            newRow.append(subRow)
                    else:
                        rows = row[0].split("\t")
                        for subRow in rows:
                            newRow.append(subRow)
                    filewriter.writerow(newRow)
            except Exception as e:
                print(e)
                pass

def convertToCSV(input_file):
    masterRows = []
    with open(input_file, "rb") as csv_in_file:
        for ln in csv_in_file:
            decoded=False
            line=''
            for cp in ('cp1252', 'cp850','utf-8','utf8'):
                try:
                    line = ln.decode(cp)
                    decoded=True
                    break
                except UnicodeDecodeError:
                    pass
            if decoded:
                try:
                    line = line.replace('0xae','')
                except:
                    pass
                try:
                    line = line.replace("├",'')
                except:
                    pass
                try:
                    line = line.replace("n┬║",'')
                except:
                    pass
                try:
                    line = line.replace("N┬║52",'')
                except:
                    pass
                try:
                    line = line.replace("1┬║D",'')
                except:
                    pass
                try:
                    line = line.replace("▒",'')
                except:
                    pass
                for i in line:
                    if not i.isalpha and not i.isnumeric():
                        line = line.replace(i,'')
                masterRows.append(line)
        with open("output.csv", "w", newline="") as csv_out_file:
            filewriter = csv.writer(csv_out_file)
            try:
                for row in masterRows:
                    indexxx = masterRows.index(row)
                    newRow = []
                    length = len(row)
                    if length > 1:
                        string = ""
                        for i in range(length):
                            string = string + row[i]
                        rows = string.split("\t")
                        for subRow in rows:
                            newRow.append(subRow)
                    else:
                        rows = row[0].split("\t")
                        for subRow in rows:
                            newRow.append(subRow)
                    filewriter.writerow(newRow)
            except Exception as e:
                print(e)
                pass
    return rows


"""
df.to_csv(
        "{d}.csv".format(d=datetime.datetime.now().strftime("%m/%d/%Y")), index=False
    )
"""

# order-id = Referencia DONE
# recipient-name = Nom. Entrega DONE
# ship-address-1 = Dirección Ent. DONE
# ship-postal-code = CPEntrega
# ship-city = Población Ent.
# ship-state = Provincia Ent.
# buyer-phone-number = Teléfono Ent.
# product-name = Observaciones 1

# IF "BAG IN BOX 15L" IN product-name and "quantity-purchase" > 1: generate new row per quantity
# IF "BAG IN BOX 5L" IN product-name and "quantity-purchase" > 2: generate new row per quantity ::: GENERATE 1 ROW PER 2 QUANTITY
# IF "PACK 2 - BAG IN BOX 15L" IN product-name and "quantity-purchase" = 1: generate new row per quantity

headers = {
    "Observaciones 2": " ",
    "Observaciones 3": " ",
    "Observaciones 4": " ",
    "Retorno": "N",
    "8:30": "N",
    "Sábado": "N",
    "Gestión/Tramite": "N",
    "Ok 15": "N",
    "Prepagado": "N",
    "Tipo Seguro": "N",
    "Imp. Seg.": "0",
    "Tipo Ealerta": " ",
    "Ealerta": " ",
    "Alto": " ",
    "Ancho": " ",
    "Largo": " ",
    "Contenido": " ",
    "Valor Declarado": " ",
    "Digitalizar/Almacenar": "N",
    "Alb.Cli.": "0",
    "Ins. Adi.": " ",
    "Ins1": " ",
    "Ins2": " ",
    "Ins3": " ",
    "Ins4": " ",
    "Ins5": " ",
    "Ins6": " ",
    "Ins7": " ",
    "Ins8": " ",
    "Ins9": " ",
    "Ins10": " ",
    "Ins11": " ",
    "Ins12": " ",
    "Ins13": " ",
    "Ins14": " ",
    "Ins15": " ",
    "Tipo PreAlerta": " ",
    "Modo Alerta": " ",
    "PreAlerta": " ",
    "PreAlerta Mensaje": " ",
    "Tipo PreAlerta1": " ",
    "Modo Alerta1": " ",
    "PreAlerta1": " ",
    "PreAlerta Mensaje1": " ",
    "Tipo PreAlerta2": " ",
    "Modo Alerta2": " ",
    "PreAlerta2": " ",
    "PreAlerta Mensaje2": " ",
    "Tipo PreAlerta3": " ",
    "Modo Alerta3": " ",
    "PreAlerta3": " ",
    "PreAlerta Mensaje3": " ",
    "Tipo PreAlerta4": " ",
    "Modo Alerta4": " ",
    "PreAlerta4": " ",
    "PreAlerta Mensaje4": " ",
    "Horario concertado": " ",
    "incio horario concertado": " ",
    "fin horario concertado": " ",
    "referencias multiples": " ",
}


def readCSV(outputFile, inputFile):
    columnsToDelete = []
    try:
        df = pd.read_csv(outputFile)
    except:
        question = input(Fore.YELLOW + "Se detectaron errores en el archivo, quisiera obtener el mayor número de filas posibles?\nDe lo contrario se obtendran todas las filas pero con otro formato. (y/n) " + Fore.RESET)
        if question == "y":
            convertAFTERQUESTION(inputFile)
            df = pd.read_csv(outputFile)
        else:
            df = pd.read_csv(outputFile, encoding="unicode_escape")
    for column in df.columns:
        if column not in columnNames:
            df.pop(column)
    formatCSV(df)


def formatCSV(df: pd.DataFrame):
    for row in df["ship-postal-code"]:
        newRow = str(row)
        while len(str(newRow)) < 5:
            newRow = "0" + str(newRow)
            rowIndex = df.loc[df["ship-postal-code"] == row].index[0]
            df["ship-postal-code"].iat[rowIndex.numerator] = newRow
    for row in df["buyer-phone-number"]:
        try:
            updatedRow  = row.replace(" ", "")
        except:
            updatedRow = row
        newRow = str(updatedRow)
        while len(str(newRow)) > 9:
            newRow = str(newRow.replace(newRow[0], "",1))
        rowIndex = df.loc[df["buyer-phone-number"] == row].index[0]
        df["buyer-phone-number"].iat[rowIndex.numerator] = newRow
    for row in df["ship-address-1"]:
        newRow = str(row)
        rowIndex = df.loc[df["ship-address-1"] == row].index[0]
        infoAddress2 = df["ship-address-2"].iat[rowIndex.numerator]
        if str(infoAddress2) == "nan":
            infoAddress2 = ""
        df["ship-address-1"].iat[rowIndex.numerator] = newRow + " " + str(infoAddress2)
    df.pop("ship-address-2")
    with ExcelWriter("outputExcel.xlsx") as writer:
        df.to_excel(writer, sheet_name="Sheet1", index=False)
        workbook: ExcelWriter = writer.book
        worksheet: ExcelWriter = writer.sheets["Sheet1"]
        for col_cells in worksheet.iter_cols(min_col=10, max_col=10):
            for cell in col_cells:
                while len(str(cell.value)) < 5:
                    cell.number_format = "@"
                    cell.value = "0" + str(cell.value)
                    writer.save()
        writer.save()
    addNewColumns(df)
    """
    df.to_csv(
        "{d}.csv".format(d=datetime.datetime.now().strftime("%m-%d-%Y")), index=False
    )
    """


def insertRow(row_number, df, row_value):
    start_upper = 0

    end_upper = row_number

    start_lower = row_number

    end_lower = df.shape[0]

    upper_half = [*range(start_upper, end_upper, 1)]

    lower_half = [*range(start_lower, end_lower, 1)]

    lower_half = [x.__add__(1) for x in lower_half]

    index_ = upper_half + lower_half

    df.index = index_

    df.loc[row_number] = row_value

    df = df.sort_index()
    # print(df)
    return df

def isALPHA(string):
    status = string.isalpha()
    return status

def formatNAMES(row):
    splitName = row.split(" ")
    finalName = {}
    for i in splitName:
        if not isALPHA(i):
            try:
                newRow = i.replace("Ã", "i")
                if isALPHA(newRow):
                    if newRow not in finalName.values():
                        finalName[splitName.index(i)] = newRow
            except:
                pass
            try:
                newRow = i.replace("Ô", "o")
                if isALPHA(newRow):
                    if newRow not in finalName.values():
                        finalName[splitName.index(i)] = newRow
            except:
                pass
            try:
                if isALPHA(newRow):
                    if newRow not in finalName.values():
                        finalName[splitName.index(i)] = newRow
                else:
                    newRow = i.replace("Ã¡", "a")
            except:
                pass
            try:
                if isALPHA(newRow):
                    if newRow not in finalName.values():
                        finalName[splitName.index(i)] = newRow
                else:
                    newRow = i.replace("Ã©", "e")
            except:
                pass
            try:
                if isALPHA(newRow):
                    if newRow not in finalName.values():
                        finalName[splitName.index(i)] = newRow
                else:
                    newRow = i.replace("Ã±", "n")
            except:
                pass
            try:
                if isALPHA(newRow):
                    if newRow not in finalName.values():
                        finalName[splitName.index(i)] = newRow
                else:
                    newRow = i.replace("Ã³", "o")
            except:
                pass
            try:
                if isALPHA(newRow):
                    if newRow not in finalName.values():
                        finalName[splitName.index(i)] = newRow
                else:
                    newRow = i.replace("Ãº", "u")
            except:
                pass
        elif splitName.index(i) == len(splitName) - 1:
            for index in finalName:
                splitName[index] = finalName[index]
    return " ".join(splitName)

def addNewColumns(df: pd.DataFrame):
    """with ExcelWriter("outputExcel.xlsx") as writer:
    df.to_excel(writer, sheet_name="Sheet1")
    workbook: ExcelWriter = writer.book
    worksheet: ExcelWriter = writer.sheets["Sheet1"]"""

    df = df.rename(
        index=str,
        columns={
            "order-id": "Referencia",
            "recipient-name": "Nom. Entrega",
            "ship-address-1": "Dirección Ent.",
            "ship-postal-code": "CPEntrega",
            "ship-city": "Población Ent.",
            "ship-state": "Provincia Ent.",
            "buyer-phone-number": "Teléfono Ent.",
            "product-name": "Observaciones 1",
        },
    )
    lengthOfRows = len(df.index)
    numbers = []
    for i in range(1, lengthOfRows + 1):
        numbers.append("37")
    df.insert(0, "Nº Abonado", numbers)

    departamentos = []
    for i in range(1, lengthOfRows + 1):
        departamentos.append(" ")
    df.insert(1, "Departamento", departamentos)

    servicios = []
    for i in range(1, lengthOfRows + 1):
        servicios.append("5")
    df.insert(2, "Servicio", servicios)

    tipoCobro = []
    for i in range(1, lengthOfRows + 1):
        tipoCobro.append("O")
    df.insert(3, "Tipo Cobro", tipoCobro)

    excesos = []
    for i in range(1, lengthOfRows + 1):
        excesos.append(" ")
    df.insert(4, "Excesos", excesos)

    bagsAndPacks = []
    for i in range(1, lengthOfRows + 1):
        bagsAndPacks.append("1")
    df.insert(6, "Bag/Paq", bagsAndPacks)

    bultos = []
    for i in range(1, lengthOfRows + 1):
        bultos.append("1")
    df.insert(7, "Bultos", bultos)

    kilos = []
    for i in range(1, lengthOfRows + 1):
        kilos.append("1")
    df.insert(8, "Kilos", kilos)

    nomEntrega = df.pop("Nom. Entrega")
    df.insert(9, "Nom. Entrega", nomEntrega)

    departamentoEnt = []
    for i in range(1, lengthOfRows + 1):
        departamentoEnt.append(" ")
    df.insert(10, "Departamento Ent.", departamentoEnt)

    personaEnt = []
    for i in range(1, lengthOfRows + 1):
        personaEnt.append(" ")
    df.insert(11, "Persona Ent.", personaEnt)

    direccionEnt = df.pop("Dirección Ent.")
    df.insert(12, "Dirección Ent.", direccionEnt)

    paisEnt = []
    for i in range(1, lengthOfRows + 1):
        paisEnt.append("ES")
    df.insert(13, "Pais Ent.", paisEnt)

    codigoPostalEnt = df.pop("CPEntrega")
    df.insert(14, "CPEntrega", codigoPostalEnt)

    poblaEnt = df.pop("Población Ent.")
    df.insert(15, "Población Ent.", poblaEnt)

    provinciaEnt = df.pop("Provincia Ent.")
    df.insert(16, "Provincia Ent.", provinciaEnt)

    telefonoEnt = df.pop("Teléfono Ent.")
    df.insert(17, "Teléfono Ent.", telefonoEnt)

    impRee = []
    for i in range(1, lengthOfRows + 1):
        impRee.append("0")
    df.insert(18, "Imp. Ree.", impRee)

    tipoRee = []
    for i in range(1, lengthOfRows + 1):
        tipoRee.append("N")
    df.insert(19, "Tipo Ree.", tipoRee)

    observaciones = df.pop("Observaciones 1")
    df.insert(20, "Observaciones 1", observaciones)
    quantityDF = df.pop("quantity-to-ship")
    for r in range(21, 82):
        for header in headers:
            columnInfo = []
            for i in range(1, lengthOfRows + 1):
                columnInfo.append(headers[header])
            df.insert(r, header, columnInfo)
            headers.pop(header)
            break
    df.insert(80, "quantity-to-ship", quantityDF)
    df = df.rename(
        columns={
            "Tipo PreAlerta1": "Tipo PreAlerta",
            "Modo Alerta1": "Modo Alerta",
            "PreAlerta1": "PreAlerta",
            "PreAlerta Mensaje1": "PreAlerta Mensaje",
            "Tipo PreAlerta2": "Tipo PreAlerta",
            "Modo Alerta2": "Modo Alerta",
            "PreAlerta2": "PreAlerta",
            "PreAlerta Mensaje2": "PreAlerta Mensaje",
            "Tipo PreAlerta2": "Tipo PreAlerta",
            "Modo Alerta3": "Modo Alerta",
            "PreAlerta3": "PreAlerta",
            "PreAlerta Mensaje3": "PreAlerta Mensaje",
            "Tipo PreAlerta3": "Tipo PreAlerta",
            "Modo Alerta4": "Modo Alerta",
            "PreAlerta4": "PreAlerta",
            "PreAlerta Mensaje4": "PreAlerta Mensaje",
        }
    )
    indexLIST = []
    index = 0
    indexL = list(df.loc[df["quantity-to-ship"] > 1 ].index)
    for i in indexL:
        if i in indexLIST and indexL.index(i) != len(indexL) - 1:
            pass
        else:
            name = df.loc[i, "Observaciones 1"]
            if "Bag in Box" in name and "5L" in name and '15L' not in name or "5 Litros" in name:
                quantity = df.loc[i, "quantity-to-ship"]
                quantityNumber = quantity.numerator
                divisibleBy = quantityNumber % 2
                if divisibleBy == 0:
                    indexToAdd = int((quantityNumber / 2) - 1)
                    indexLIST.append(str(int(i) + indexToAdd))
                else:
                    indexLIST.append(str(int(i)))
            else:
                indexLIST.append(str(int(i) + index))
                index += 1

    for row in df["quantity-to-ship"]:
        if row == 1:
            pass
        else:
            rowIndex = indexLIST[0]
            # print(df)
            rowToDuplicate = df.iloc[int(rowIndex)]
            time.sleep(0.2)
            prodNameAtRowIndex = rowToDuplicate.get(
                "Observaciones 1"
            )  # df.iloc[str(rowIndex), "Observaciones 1"]
            if "BAG IN BOX 15L" in prodNameAtRowIndex.upper() and row > 1:
                quantityToGenerate = row - 1
                for i in range(quantityToGenerate):
                    df = insertRow(int(rowIndex) + 1, df, rowToDuplicate)

            elif "BAG IN BOX 5L" or "BAG IN BOX" and "5 LITROS" in prodNameAtRowIndex.upper():
                if row <= 2:
                    indexLISTquant = list(df.loc[df["Observaciones 1"] == prodNameAtRowIndex].index)
                    if row == 1:
                       pass
                    if row == 2:
                        newname = prodNameAtRowIndex.upper()
                        if "BLANCO" in newname and "VERDEJO" not in newname:
                            newname = "BAG IN BOX 5L BLANCO JOVEN"
                        elif "ROSADO" in newname:
                            newname = "BAG IN BOX 5L ROSADO JOVEN"
                        elif "PREMIUM" in newname:
                            newname = "BAG IN BOX 5L TINTO PREMIUM"
                        elif "RECOMENDADO" in newname:
                            newname = "BAG IN BOX 5L TINTO RECOMENDADO"
                        elif "VERDEJO" in newname and "BLANCO" not in newname:
                            newname = "BAG IN BOX 5L VERDEJO PAZ VI"
                        elif "NUEVO REINO" in newname:
                            newname = "BAG IN BOX 5L TINTO NUEVO REINO"
                        elif "VERDEJO" in newname and "BLANCO" in newname:
                            newname = "BAG IN BOX 5L BLANCO VERDEJO"
                        df.loc[
                            df["Observaciones 1"] == prodNameAtRowIndex, "Observaciones 1"
                        ] = "PACK (2) " + newname
                else:
                    remainder = row % 2
                    index = 0
                    if remainder == 0:
                        division = (row / 2) - 1
                        for i in range(int(division)):
                            df = insertRow(int(rowIndex) + 1 + index, df, rowToDuplicate)
                        listWhereProds = list(df.loc[df["Observaciones 1"] == prodNameAtRowIndex].index)
                        for i in listWhereProds:
                            newname = df.loc[i]["Observaciones 1"].upper()[:40]
                            if "BLANCO" in newname and "VERDEJO" not in newname:
                                newname = "BAG IN BOX 5L BLANCO JOVEN"
                            elif "ROSADO" in newname:
                                newname = "BAG IN BOX 5L ROSADO JOVEN"
                            elif "PREMIUM" in newname:
                                newname = "BAG IN BOX 5L TINTO PREMIUM"
                            elif "RECOMENDADO" in newname:
                                newname = "BAG IN BOX 5L TINTO RECOMENDADO"
                            elif "VERDEJO" in newname  and "BLANCO" not in newname:
                                newname = "BAG IN BOX 5L VERDEJO PAZ VI"
                            elif "NUEVO REINO" in newname:
                                newname = "BAG IN BOX 5L TINTO NUEVO REINO"
                            elif "VERDEJO" in newname and "BLANCO" in newname:
                                newname = "BAG IN BOX 5L BLANCO VERDEJO"
                            df.at[df.index[int(i)], "Observaciones 1"] = "PACK (2) " + newname
                    else:
                        newLimit = int(((row - 1) / 2) + remainder)
                        for i in range(newLimit):
                            if i < newLimit - 1:
                                df = insertRow(int(rowIndex) + 1 + index, df, rowToDuplicate)
                                listWhereProds = list(df.loc[df["Observaciones 1"] == prodNameAtRowIndex].index)
                                newname = prodNameAtRowIndex.upper()[:40]
                                if "BLANCO" in newname and "VERDEJO" not in newname:
                                    newname = "BAG IN BOX 5L BLANCO JOVEN"
                                elif "ROSADO" in newname:
                                    newname = "BAG IN BOX 5L ROSADO JOVEN"
                                elif "PREMIUM" in newname:
                                    newname = "BAG IN BOX 5L TINTO PREMIUM"
                                elif "RECOMENDADO" in newname:
                                    newname = "BAG IN BOX 5L TINTO RECOMENDADO"
                                elif "VERDEJO" in newname and "BLANCO" not in newname:
                                    newname = "BAG IN BOX 5L VERDEJO PAZ VI"
                                elif "NUEVO REINO" in newname:
                                    newname = "BAG IN BOX 5L TINTO NUEVO REINO"
                                elif "VERDEJO" in newname and "BLANCO" in newname:
                                    newname = "BAG IN BOX 5L BLANCO VERDEJO PAZ VI"
                                df.at[df.index[int(rowIndex) + 1 + index], "Observaciones 1"] = "PACK (2) " + newname
                                index += 1
                            else:
                                pass
            elif "BAG IN BOX 3L" in prodNameAtRowIndex.upper() or "BAG IN BOX" and "3 LITROS" in prodNameAtRowIndex.upper():
                if row <= 2:
                    indexLISTquant = list(df.loc[df["Observaciones 1"] == prodNameAtRowIndex].index)
                    if row == 1:
                       pass
                    if row == 2:
                        newname = prodNameAtRowIndex.upper()
                        if "RECOMENDADO" in newname:
                            newname = "BAG IN BOX 3L TINTO RECOMENDADO"
                        elif "BAG IN BOX VINO BLANCO VERDEJO 3 LITROS" in newname:
                            newname = "BAG IN BOX 3L VERDEJO PAZ VI"
                        elif "BAG IN BOX 3L VINO TINTO NUEVO REINO" in newname:
                            newname = "BAG IN BOX 3L TINTO NUEVO REINO"
                        elif "BAG IN BOX VINO BLANCO VERDEJO" in newname:
                            newname = "BAG IN BOX 3L BLANCO VERDEJO"
                        df.loc[
                            df["Observaciones 1"] == prodNameAtRowIndex, "Observaciones 1"
                        ] = "PACK (2) " + newname
                else:
                    remainder = row % 2
                    index = 0
                    if remainder == 0:
                        division = (row / 2) - 1
                        for i in range(int(division)):
                            df = insertRow(int(rowIndex) + 1 + index, df, rowToDuplicate)
                        listWhereProds = list(df.loc[df["Observaciones 1"] == prodNameAtRowIndex].index)
                        for i in listWhereProds:
                            newname = df.loc[i]["Observaciones 1"].upper()[:40]
                            if "BAG IN BOX VINO BLANCO VERDEJO 3 LITROS" in newname:
                                newname = "BAG IN BOX 3L VERDEJO PAZ VI"
                            elif "BAG IN BOX 3L VINO TINTOS RECOMENDADO" in newname:
                                newname = "BAG IN BOX 3L TINTO RECOMENDADO"
                            elif "BAG IN BOX 3L VINO TINTO NUEVO REINO" in newname:
                                newname = "BAG IN BOX 3L TINTO NUEVO REINO"
                            elif "BAG IN BOX VINO BLANCO VERDEJO" in newname:
                                newname = "BAG IN BOX 3L BLANCO VERDEJO"
                            df.at[df.index[int(i)], "Observaciones 1"] = "PACK (2) " + newname
                    else:
                        newLimit = int(((row - 1) / 2) + remainder)
                        for i in range(newLimit):
                            if i < newLimit - 1:
                                df = insertRow(int(rowIndex) + 1 + index, df, rowToDuplicate)
                                listWhereProds = list(df.loc[df["Observaciones 1"] == prodNameAtRowIndex].index)
                                newname = prodNameAtRowIndex.upper()[:40]
                                if "BAG IN BOX VINO BLANCO VERDEJO 3 LITROS" in newname:
                                    newname = "BAG IN BOX 3L VERDEJO PAZ VI"
                                elif "BAG IN BOX 3L VINO TINTOS RECOMENDADO" in newname:
                                    newname = "BAG IN BOX 3L TINTO RECOMENDADO"
                                elif "BAG IN BOX 3L VINO TINTO NUEVO REINO" in newname:
                                    newname = "BAG IN BOX 3L TINTO NUEVO REINO"
                                elif "BAG IN BOX VINO BLANCO VERDEJO" in newname:
                                    newname = "BAG IN BOX 3L BLANCO VERDEJO"
                                df.at[df.index[int(rowIndex) + 1 + index], "Observaciones 1"] = "PACK (2) " + newname
                                index += 1
                            else:
                                pass
            indexLIST.pop(indexLIST.index(rowIndex))

    for row in df["Observaciones 1"]:
        if "PACK - 2" in row:
            index = 0
            indexes = list(df.loc[df["Observaciones 1"] == row].index)
            quantity = list(df.loc[df["Observaciones 1"] == row, "quantity-to-ship"])
            for r in range(len(indexes)):
                nameAT = df.iloc[int(indexes[r]) + index]["Nom. Entrega"]
                #("Name at {i} is {name}".format(i=indexes[r] + index, name=nameAT))
                #print('Name at 21' + str(df.iloc[int(21)]["Nom. Entrega"]))
                #print('Name at 22' + str(df.iloc[int(22)]["Nom. Entrega"]))
                #print('Name at 23' + str(df.iloc[int(23)]["Nom. Entrega"]))
                amountToDuplicate = (quantity[r] * 2) - 1
                for i in range(amountToDuplicate):
                    rowToDuplicate = df.iloc[int(indexes[r]) + index]
                    df = insertRow(int(indexes[r]) + 1 + index, df, rowToDuplicate)
                    index += 1
            break
    
    df.pop("quantity-to-ship")
    
    for row in df["Nom. Entrega"]:
        newRow = formatNAMES(row)
        df.loc[df["Nom. Entrega"] == row, "Nom. Entrega"] = newRow.upper()

    for row in df["Observaciones 1"]:
        if "Bag in Box 15L Vino Tinto" in row:
            if "Recomendado" in row:
                df.loc[
                    df["Observaciones 1"] == row, "Observaciones 1"
                ] = "BAG IN BOX 15L TINTO RECOMENDADO"
            elif "PREMIUM" in row:
                df.loc[
                    df["Observaciones 1"] == row, "Observaciones 1"
                ] = "BAG IN BOX 15L TINTO PREMIUM"
            elif "Joven" in row:
                df.loc[
                    df["Observaciones 1"] == row, "Observaciones 1"
                ] = "BAG IN BOX 15L TINTO JOVEN"
        elif "PACK - 2" in row:
            df.loc[
                df["Observaciones 1"] == row, "Observaciones 1"
            ] = "BAG IN BOX 15L TINTO CAJA BARRICA"
        elif "Bag in Box 5L" in row or "BAG IN BOX 5L" in row or "BAG IN BOX" and "5 LITROS" in row:
            if "PREMIUM" in row:
                df.loc[
                    df["Observaciones 1"] == row, "Observaciones 1"
                ] = "BAG IN BOX 5L TINTO PREMIUM"
            elif "Recomendado" in row:
                df.loc[
                    df["Observaciones 1"] == row, "Observaciones 1"
                ] = "BAG IN BOX 5L TINTO RECOMENDADO"
            elif "cosechero" in row and "blanco" not in row:
                df.loc[
                    df["Observaciones 1"] == row, "Observaciones 1"
                ] = "BAG IN BOX 5L TINTO CAJA BARRICA"
            elif "BLANCO" in row.upper() and "PACK (2)" not in row:
                df.loc[
                    df["Observaciones 1"] == row, "Observaciones 1"
                ] = "BAG IN BOX 5L BLANCO JOVEN"
            elif "ROSADO" in row.upper() and "PACK (2)" not in row:
                df.loc[
                    df["Observaciones 1"] == row, "Observaciones 1"
                ] = "BAG IN BOX 5L ROSADO JOVEN"
            elif "Joven" in row:
                df.loc[
                    df["Observaciones 1"] == row, "Observaciones 1"
                ] = "BAG IN BOX 5L TINTO JOVEN"
            elif "Verdejo" in row and "Blanco" in row:
                df.loc[
                    df["Observaciones 1"] == row, "Observaciones 1"
                ] = "BAG IN BOX 5L VERDEJO PAZ VI"
        elif "Bag in Box verdejo 15 Litros" in row:
            df.loc[
                df["Observaciones 1"] == row, "Observaciones 1"
            ] = "BAG IN BOX 15L VERDEJO PAZ VI"
        elif "Bag in Box 15L Vino" in row:
            if "cosechero" in row:
                df.loc[
                    df["Observaciones 1"] == row, "Observaciones 1"
                ] = "BAG IN BOX 15L TINTO CAJA BARRICA"
            elif "Blanco" in row:
                df.loc[
                    df["Observaciones 1"] == row, "Observaciones 1"
                ] = "BAG IN BOX 15L BLANCO JOVEN"
            elif "Rosado" in row:
                df.loc[
                    df["Observaciones 1"] == row, "Observaciones 1"
                ] = "BAG IN BOX 15L ROSADO JOVEN"
            elif "Joven" in row:
                df.loc[
                    df["Observaciones 1"] == row, "Observaciones 1"
                ] = "BAG IN BOX 15L TINTO JOVEN"
        elif "BOTELLAS 6" in row.upper() or "6 BOTELLAS" in row.upper():
            if "PAULUS JOVEN" in row.upper():
                df.loc[
                    df["Observaciones 1"] == row, "Observaciones 1"
                ] = "C-6 BOTELLAS PAULUS JOVEN RIOJA"
            elif "PAULUS CRIANZA" in row.upper():
                df.loc[
                    df["Observaciones 1"] == row, "Observaciones 1"
                ] = "C-6 BOTELLAS PAULUS CRIANZA RIOJA"  
            elif "SOTONOVILLOS CRIANZA" in row.upper():
                df.loc[
                    df["Observaciones 1"] == row, "Observaciones 1"
                ] = "C-6 BOTELLAS SOTONOVILLOS CRIANZA RIOJA"
            elif "SIDRA NATURAL" in row.upper():
                df.loc[
                    df["Observaciones 1"] == row, "Observaciones 1"
                ] = "C-6 BOTELLAS SIDRA NATURAL JAREGUI"
            elif "FAUNA IBERICA" in row.upper():
                df.loc[
                    df["Observaciones 1"] == row, "Observaciones 1"
                ] = "C-6 BOTELLAS COLECCION FAUNA IBERICA"
            elif "PANJUA" in row.upper():
                df.loc[
                    df["Observaciones 1"] == row, "Observaciones 1"
                ] = "C-6 BOTELLAS SEÑORÍO DE PANJUA VERDEJO"
            elif "PREMIUM VINO TINTO" and "BOTELLAS 6" in row.upper() or "6 BOTELLAS" in row.upper():
                df.loc[
                    df["Observaciones 1"] == row, "Observaciones 1"
                ] = "C-6 BOTELLAS TINTO PREMIUM LOS CORZOS"
            elif "CLARETE" and "BOTELLAS 6" in row.upper() or "6 BOTELLAS" in row.upper():
                df.loc[
                    df["Observaciones 1"] == row, "Observaciones 1"
                ] = "C-6 BOTELLAS CLARETE LOS CORZOS"
            elif "COSECHERO" and "BOTELLAS 6" in row.upper() or "6 BOTELLAS" in row.upper():
                df.loc[
                    df["Observaciones 1"] == row, "Observaciones 1"
                ] = "C-6 BOTELLAS TINTO LOS CORZOS"
            elif "BLANCO" and "BOTELLAS 6" in row.upper() or "6 BOTELLAS" in row.upper():
                df.loc[
                    df["Observaciones 1"] == row, "Observaciones 1"
                ] = "C-6 BOTELLAS BLANCO LOS CORZOS"
            elif "ROSADO" and "BOTELLAS 6" in row.upper() or "6 BOTELLAS" in row.upper():
                df.loc[
                    df["Observaciones 1"] == row, "Observaciones 1"
                ] = "C-6 BOTELLAS ROSADO LOS CORZOS"
            elif "VERDEJO" and "BOTELLAS 6" in row.upper() or "6 BOTELLAS" in row.upper():
                df.loc[
                    df["Observaciones 1"] == row, "Observaciones 1"
                ] = "C-6 BOTELLAS VERDEJO PAZ VI"
            elif "TINTO RECOMENDADO LOS CORZOS" and "BOTELLAS 6" in row.upper() or "6 BOTELLAS" in row.upper():
                df.loc[
                    df["Observaciones 1"] == row, "Observaciones 1"
                ] = "C-6 BOTELLAS TINTO RECOMENDADO  B/N LOS CORZOS"
            elif "EL APARATO VINO BLANCO" and "BOTELLAS 6" in row.upper() or "6 BOTELLAS" in row.upper():
                df.loc[
                    df["Observaciones 1"] == row, "Observaciones 1"
                ] = "C-6 BOTELLAS TE ENSEÑO EL APARATO"
            elif "NUEVO REINO TINTO" and "BOTELLAS 6" in row.upper() or "6 BOTELLAS" in row.upper():
                df.loc[
                    df["Observaciones 1"] == row, "Observaciones 1"
                ] = "C-6 BOTELLAS TINTO NUEVO REINO"
            elif "RUFUS" and "BOTELLAS 6" in row.upper() or "6 BOTELLAS" in row.upper():
                df.loc[
                    df["Observaciones 1"] == row, "Observaciones 1"
                ] = "C-6 BOTELLAS TINTO RUFUS RIBERA DUERO"
        elif "BAG IN BOX 3L" in row.upper() or "BAG IN BOX" and "3 LITROS" in row.upper():
            if "COSECHERO" in row.upper():
                df.loc[
                    df["Observaciones 1"] == row, "Observaciones 1"
                ] = "BAG IN BOX 3L TINTO CAJA BARRICA"
            elif "RECOMENDADO" in row.upper():
                df.loc[
                    df["Observaciones 1"] == row, "Observaciones 1"
                ] = "BAG IN BOX 3L TINTO RECOMENDADO"
            elif "VERDEJO" in row.upper() and "BLANCO" in row.upper():
                df.loc[
                    df["Observaciones 1"] == row, "Observaciones 1"
                ] = "BAG IN BOX 3L BLANCO VERDEJO AFRUTADO"
    with ExcelWriter(
        "{d}.xlsx".format(d=datetime.datetime.now().strftime("%d-%m-%Y"))
    ) as writer:
        df.to_excel(writer, sheet_name="Sheet1", index=False)
        workbook: ExcelWriter = writer.book
        worksheet: ExcelWriter = writer.sheets["Sheet1"]
        for col_cells in worksheet.iter_cols(min_col=15, max_col=15):
            for cell in col_cells:
                while len(str(cell.value)) < 5:
                    cell.number_format = "@"
                    cell.value = "0" + str(cell.value)
                    writer.save()
        writer.save()
    # print(df)

def main():
    rawStringPath = input("Introduzca la direccion del archivo: ")
    rawStringPath = rawStringPath.replace('"', "")
    convertToCSV(r"{s}".format(s=rawStringPath))
    readCSV(r"{s}\output.csv".format(s=os.getcwd()),r"{s}".format(s=rawStringPath))
    filename = "{d}.xlsx".format(d=datetime.datetime.now().strftime("%d-%m-%Y"))
    base = os.path.splitext(filename)[0]
    
    os.remove(r"{s}\output.csv".format(s=os.getcwd()))
    os.remove(r"{s}\outputExcel.xlsx".format(s=os.getcwd()))
    os.system("cls")
    print(
        Fore.GREEN
        + 'Proceso finalizado.\nArchivo generado: "{f}"'.format(f=filename)
        + Fore.RESET
    )
    time.sleep(5)


if __name__ == "__main__":
    main()
